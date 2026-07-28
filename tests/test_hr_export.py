"""Tests for HR export endpoints."""

import csv
from io import BytesIO, StringIO

import pytest


async def test_export_shifts_requires_admin(client):
    resp = await client.get(
        "/api/v1/hr/exports/shifts",
        headers={"Authorization": "Bearer invalid-token"},
    )
    assert resp.status_code in (401, 403)


async def test_export_timeclock_requires_admin(client):
    resp = await client.get(
        "/api/v1/hr/exports/timeclock",
        headers={"Authorization": "Bearer invalid-token"},
    )
    assert resp.status_code in (401, 403)


@pytest.fixture
async def default_establishment_id(db_session):
    import sqlalchemy as sa

    await db_session.execute(sa.text('SET search_path TO "tenant_pizza_test", public'))
    result = await db_session.execute(
        sa.text("SELECT id FROM establishments ORDER BY id LIMIT 1")
    )
    establishment_id = result.scalar_one()
    result.close()
    return establishment_id


async def test_export_shifts_csv_contains_header_and_row(
    db_session,
    default_establishment_id,
):
    from app.modules.hr.schemas import EmployeeProfileCreate, ShiftCreate
    from app.modules.hr.service import (
        create_employee_profile,
        create_shift,
        export_shifts_csv,
    )

    employee = await create_employee_profile(
        db_session,
        EmployeeProfileCreate(
            user_id=99008,
            establishment_id=default_establishment_id,
        ),
    )
    await create_shift(
        db_session,
        ShiftCreate(
            employee_id=employee.id,
            establishment_id=default_establishment_id,
            starts_at="2026-08-01T08:00:00Z",
            ends_at="2026-08-01T16:00:00Z",
        ),
        created_by_user_id=1,
    )

    csv_text = await export_shifts_csv(db_session, employee_id=employee.id)
    rows = list(csv.reader(StringIO(csv_text)))
    assert rows[0] == ["id", "employee_id", "establishment_id", "starts_at", "ends_at", "status"]
    assert len(rows) == 2


async def test_export_time_clock_csv_contains_header_and_row(
    db_session,
    default_establishment_id,
):
    from app.modules.hr.schemas import ClockInRequest, EmployeeProfileCreate
    from app.modules.hr.service import (
        clock_in,
        clock_out,
        create_employee_profile,
        export_time_clock_csv,
    )

    employee = await create_employee_profile(
        db_session,
        EmployeeProfileCreate(
            user_id=99009,
            establishment_id=default_establishment_id,
        ),
    )
    await clock_in(
        db_session,
        employee.id,
        ClockInRequest(method="web", establishment_id=default_establishment_id),
    )
    await clock_out(db_session, employee.id)

    csv_text = await export_time_clock_csv(db_session, employee_id=employee.id)
    rows = list(csv.reader(StringIO(csv_text)))
    assert rows[0] == [
        "id",
        "employee_id",
        "clock_in_at",
        "clock_out_at",
        "method",
        "status",
    ]
    assert len(rows) == 2


async def test_export_shifts_xlsx_returns_valid_workbook(
    db_session,
    default_establishment_id,
):
    import openpyxl

    from app.modules.hr.schemas import EmployeeProfileCreate, ShiftCreate
    from app.modules.hr.service import (
        create_employee_profile,
        create_shift,
        export_shifts_xlsx,
    )

    employee = await create_employee_profile(
        db_session,
        EmployeeProfileCreate(
            user_id=99010,
            establishment_id=default_establishment_id,
        ),
    )
    await create_shift(
        db_session,
        ShiftCreate(
            employee_id=employee.id,
            establishment_id=default_establishment_id,
            starts_at="2026-08-01T08:00:00Z",
            ends_at="2026-08-01T16:00:00Z",
        ),
        created_by_user_id=1,
    )

    content = await export_shifts_xlsx(db_session, employee_id=employee.id)
    workbook = openpyxl.load_workbook(BytesIO(content))
    sheet = workbook.active
    assert sheet["A1"].value == "id"
    assert sheet.max_row == 2


async def test_export_shifts_pdf_returns_pdf_bytes(
    db_session,
    default_establishment_id,
):
    from app.modules.hr.schemas import EmployeeProfileCreate, ShiftCreate
    from app.modules.hr.service import (
        create_employee_profile,
        create_shift,
        export_shifts_pdf,
    )

    employee = await create_employee_profile(
        db_session,
        EmployeeProfileCreate(
            user_id=99011,
            establishment_id=default_establishment_id,
        ),
    )
    await create_shift(
        db_session,
        ShiftCreate(
            employee_id=employee.id,
            establishment_id=default_establishment_id,
            starts_at="2026-08-01T08:00:00Z",
            ends_at="2026-08-01T16:00:00Z",
        ),
        created_by_user_id=1,
    )

    content = await export_shifts_pdf(db_session, employee_id=employee.id)
    assert content[:5] == b"%PDF-"


async def test_export_time_clock_xlsx_returns_valid_workbook(
    db_session,
    default_establishment_id,
):
    import openpyxl

    from app.modules.hr.schemas import ClockInRequest, EmployeeProfileCreate
    from app.modules.hr.service import (
        clock_in,
        clock_out,
        create_employee_profile,
        export_time_clock_xlsx,
    )

    employee = await create_employee_profile(
        db_session,
        EmployeeProfileCreate(
            user_id=99012,
            establishment_id=default_establishment_id,
        ),
    )
    await clock_in(
        db_session,
        employee.id,
        ClockInRequest(method="web", establishment_id=default_establishment_id),
    )
    await clock_out(db_session, employee.id)

    content = await export_time_clock_xlsx(db_session, employee_id=employee.id)
    workbook = openpyxl.load_workbook(BytesIO(content))
    sheet = workbook.active
    assert sheet["A1"].value == "id"
    assert sheet.max_row == 2


async def test_export_time_clock_pdf_returns_pdf_bytes(
    db_session,
    default_establishment_id,
):
    from app.modules.hr.schemas import ClockInRequest, EmployeeProfileCreate
    from app.modules.hr.service import (
        clock_in,
        clock_out,
        create_employee_profile,
        export_time_clock_pdf,
    )

    employee = await create_employee_profile(
        db_session,
        EmployeeProfileCreate(
            user_id=99013,
            establishment_id=default_establishment_id,
        ),
    )
    await clock_in(
        db_session,
        employee.id,
        ClockInRequest(method="web", establishment_id=default_establishment_id),
    )
    await clock_out(db_session, employee.id)

    content = await export_time_clock_pdf(db_session, employee_id=employee.id)
    assert content[:5] == b"%PDF-"
